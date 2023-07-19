import pyautogui
import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import shutil
from pathlib import Path
import zipfile
import glob
import csv
import openpyxl as px
from openpyxl.utils import column_index_from_string
import sys

data = ['293212000', '293213000']

# 一括削除関数
def remove_glob(pathname, recursive=True):
    for p in glob.glob(pathname, recursive=recursive):
        if os.path.isfile(p):
            os.remove(p)

def dldata():
    # ここから目的に合わせて今回はfunc1-func4で動作させる。以下はfunc2の例
    # 適宜データ用意し、同様に処理
    for index, item in enumerate(data):
        print(index, item)
        print("Call func1")
        if index == 0:
            func1(item,data)
        elif index == 1:
            func2(item,data)

    # DL OPEN
    pyautogui.sleep(5)
    pyautogui.click(x=100, y=1050)

def output(data,i):
    print("[debug: ]",data)

    # 
    # 
    # Output -> CSV
    # 
    # 
    time.sleep(5)
    pyautogui.click(x=40, y=420)
    for i in range(1):
        pyautogui.press('tab')
    time.sleep(1)
    pyautogui.press('enter')
    print("DL......................")
    time.sleep(1)

    if i == 1:
        pyautogui.keyDown('shift')
        for i in range(2):
            pyautogui.press('tab')
        pyautogui.keyUp('shift')
    pyautogui.press('enter')
    print("-------------------再検索-------------------")

# func1: 特定の国コードを指定して検索し、CSVをダウンロードする
def func1(n,data):

    # 
    # 
    # SearchPage(https://www.customs.go.jp/toukei/srch/index.htm?M=77&P=...)
    # 
    # 

    pyautogui.sleep(6)
    # pyautogui.click(x=40, y=375)
    pyautogui.sleep(1)
    # pyautogui.press('tab')
    pyautogui.press('right')
    print("輸入")

    for i in range(2):
        pyautogui.press('tab')
    time.sleep(1)
    pyautogui.press('space')
    time.sleep(1)
    for i in range(5):
        pyautogui.press('down')
    time.sleep(1)
    pyautogui.press('enter')
    print("get year")

    for i in range(3):
        pyautogui.press('tab')
    pyautogui.press('space')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.press('tab')
    time.sleep(1)
    # 105:China
    pyautogui.typewrite('105')
    # time.sleep(1)
    for i in range(3):
        pyautogui.press('tab')
    print("get country")

    pyautogui.press('space')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.press('tab')
    pyautogui.typewrite(n)
    # time.sleep(1)
    pyautogui.press('enter')

    print("TestFinished")
    i=1
    output(data,i)

def func2(n,data):
    time.sleep(4)
    print("Select code...........")
    for i in range(10):
        pyautogui.press('tab')
    # time.sleep(1)
    for i in range(12):
        pyautogui.press('backspace')
    # time.sleep(1)
    pyautogui.typewrite(n)
    pyautogui.press('enter')

    print("goto OutputPage")
    i=2
    output(data,i)

# func3: 国コードと品目コードを指定して検索し、CSVをダウンロードする
def func3(country_num, n):
    # TAB+ALT
    pyautogui.sleep(5)
    pyautogui.keyDown('alt')
    pyautogui.press('tab')
    pyautogui.keyUp('alt')

    pyautogui.sleep(1)
    pyautogui.click(x=40, y=375)

    # Number INPUT
    pyautogui.sleep(2)
    for i in range(6):
        pyautogui.press('tab')
    pyautogui.typewrite(country_num)

    for i in range(4):
        pyautogui.press('tab')

    # Type Keyboard
    pyautogui.typewrite(n)

    # Search
    pyautogui.press('enter')

    # CSV DL
    pyautogui.sleep(3)
    pyautogui.click(x=70, y=460)

# func4: 品目コードを指定して検索し、CSVをダウンロードする
def func4(n):
    # TAB+ALT
    pyautogui.sleep(5)
    pyautogui.keyDown('alt')
    pyautogui.press('tab')
    pyautogui.keyUp('alt')

    pyautogui.sleep(1)
    pyautogui.click(x=40, y=375)

    # Number INPUT
    # Select All Country
    pyautogui.sleep(2)
    for i in range(5):
        pyautogui.press('tab')
    pyautogui.press('up')
    pyautogui.press('tab')
    pyautogui.press('tab')

    # Type Keyboard
    pyautogui.press('delete')
    pyautogui.typewrite(n)

    # Search
    pyautogui.press('enter')

    # CSV DL
    pyautogui.sleep(3)
    pyautogui.click(x=70, y=460)

# ダウンロード関数
def exebrowser():
    # run chrome
    driver = webdriver.Chrome(ChromeDriverManager().install())

    driver.get('https://www.customs.go.jp/toukei/srch/index.htm?M=77&P=0')
    driver.maximize_window()

    # dlデータ関数へ
    dldata()
    driver.close()

try:
    # ホームディレクトリまでのパスを取得
    p_path = os.path.expanduser("~")
    remove_glob(p_path + 'Downloads/*.zip')
    x, y = pyautogui.position()

    # exebrowser関数の実行
    exebrowser()

    # ダウンロードファイルを一時フォルダに解凍し、リネームする
    today = datetime.date.today()  # 今日の日付を取得
    Y = str(today.year)
    M = str(today.month)
    D = str(today.day)

    # Path形式でDownloads以下のzipファイルを l_1にlistで格納
    p = Path(p_path + "Downloads/")
    l_1 = list(p.glob('*.zip'))
    re_l = ['fl.csv', 'fa.csv']

    Temp_P = Path(p_path + "/DESKTOP/temp_python")

    # listの要素数から、zipfile.ZipFileリストをmyzipに格納
    for i in range(len(l_1)):
        with zipfile.ZipFile(str(l_1[i])) as myzip:
            # Temp_Pフォルダへ取り出し
            myzip.extractall(Temp_P)

    # temp_pythonフォルダの中のファイルを取得し、リネームして保存する
    files = sorted(Temp_P.glob('*.csv'))
    i = 0

    # 格納先フォルダの作成
    NEW_P = Path(p_path + "/DESKTOP/temp_python/RENAME(PY_AUTO)/")
    if not os.path.exists(NEW_P):
        os.mkdir(NEW_P)
    else:
        shutil.rmtree(NEW_P)
        time.sleep(1)
        os.mkdir(NEW_P)

    # ファイルのリネーム
    for file in files:
        # 隠しファイルは処理しない
        if file.name.startswith('.'):
            continue

        # old Path
        oldFilePath = file.__str__()
        print("oldFilePath:", oldFilePath)

        # new Path
        newFileName = re_l[i]
        newFilePath = str(NEW_P) + '/' + newFileName
        print("newFileName:", newFileName)
        print("newFilePath:", newFilePath)

        os.rename(oldFilePath, newFilePath)
        i += 1

    # CSVをEXCELに変換し、データをコピー＆ペーストする
    with open(p_path + 'Desktop/temp_python/RENAME(PY_AUTO)/.csv', newline='') as tempfile:
        data_list_fl = list(csv.reader(tempfile, delimiter=','))

    wb = px.Workbook()
    ws = wb.active
    ws.append([])  # A1, B1 空行

    # データの格納
    for record in data_list_fl:
        ws.append(record)

    # xlsx形式で保存
    wb.save(p_path + 'Desktop/temp_python/RENAME(PY_AUTO)/temp.xlsx')

    # データを別のブックにコピー＆ペースト
    temp = p_path + 'Desktop/temp_python/RENAME(PY_AUTO)/temp.xlsx'
    output = p_path + 'Desktop/test.xlsx'

    wb_temp = px.load_workbook(filename=temp)
    ws_temp = wb_temp.active

    wb_output = px.Workbook()
    ws_output = wb_output.active

    for row in ws_temp.iter_rows(values_only=True):
        ws_output.append(row)

    # データを保存
    wb_output.save(output)

except KeyboardInterrupt:
    print('\n終了')
except Exception as e:
    print('エラーが発生しました:', str(e))
    # エラーメッセージを表示したり、適切な処理を行ったりする
