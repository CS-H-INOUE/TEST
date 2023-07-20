import os
import csv
import openpyxl as px
import sys

def get_user_desktop_path():
    """ユーザーのデスクトップのパスを取得する関数"""
    return os.path.join(os.path.expanduser('~'), 'Desktop')

def get_sheet_name(filename):
    """ファイル名（.csvを除く）を取得してシート名に設定する関数"""
    return os.path.splitext(filename)[0]

def clear_sheet(ws):
    """指定されたシートのすべてのセルを削除する関数"""
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)

def copy_csv_data_to_excel(csv_file_path, ws_output, debug_mode=False):
    """CSVファイルのデータをExcelシートにコピーする関数"""
    with open(csv_file_path, 'r', encoding='shift-jis', newline='') as csvfile:
        data_list_fl = list(csv.reader(csvfile, delimiter=','))

    for record in data_list_fl:
        ws_output.append(record)
        # デバッグモードであればデータを追加する際にプリント
        if debug_mode:
            print(f"データを追加: {record}")

def main():
    # ユーザーデスクトップのパスを取得
    desktop_path = get_user_desktop_path()

    # フォルダパスと出力ファイルパスを設定
    folder_path = os.path.join(desktop_path, 'temp_python/RENAME(PY_AUTO)')
    output_file = os.path.join(desktop_path, 'PythonScripts/Code/AutoGUI/FLFA/FL,FADATA.xlsx')

    # 引数に"--debug"がある場合はデバッグモードとして実行
    debug_mode = "--debug" in sys.argv

    # 既存のExcelファイルを開く
    wb_output = px.load_workbook(filename=output_file)

    # デバッグ用のプリント文を追加
    print("フォルダパス:", folder_path)
    print("出力ファイル:", output_file)

    for filename in os.listdir(folder_path):
        if filename.endswith('.csv'):
            csv_file_path = os.path.join(folder_path, filename)
            sheet_name = get_sheet_name(os.path.basename(csv_file_path))

            # シートが既に存在する場合は取得し、ない場合は新規作成
            if sheet_name in wb_output.sheetnames:
                ws_output = wb_output[sheet_name]
                # 既存のシート内のすべてのセルの内容を削除
                clear_sheet(ws_output)
            else:
                ws_output = wb_output.create_sheet(title=sheet_name)

            copy_csv_data_to_excel(csv_file_path, ws_output, debug_mode)

    # デバッグモードの場合のみプリント
    if debug_mode:
        print("データをコピー完了")

    # 保存して終了
    wb_output.save(output_file)

    # デバッグモードの場合のみプリント
    if debug_mode:
        print("ファイルを保存完了")

if __name__ == "__main__":
    main()
