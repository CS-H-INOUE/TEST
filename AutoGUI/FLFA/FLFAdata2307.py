import pyautogui
import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import shutil
from pathlib import Path
import zipfile
import glob
import csv
import openpyxl as px
from openpyxl.utils import column_index_from_string
import sys

# 一括削除関数
def remove_glob(pathname, recursive=True):
    for p in glob.glob(pathname, recursive=recursive):
        if os.path.isfile(p):
            os.remove(p)

data = ['293212000', '293213000']

# ダウンロードデータの処理関数
def dldata(data):
    # Chromeを起動しWebページにアクセス
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get('https://www.customs.go.jp/toukei/srch/index.htm?M=77&P=0')
    driver.maximize_window()

    # ダウンロードデータ処理
    for index, item in enumerate(data):
        print(index, item)
        if index == 0:
            func1(item, data)
        elif index == 1:
            func2(item, data)

    # DL OPEN
    pyautogui.sleep(5)
    pyautogui.click(x=100, y=1050)
    driver.close()

# 出力処理
def output(data, i):
    print("[debug: ]", data)

    # CSVをEXCELに変換
    time.sleep(5)
    pyautogui.click(x=40, y=420)
    for _ in range(1):
        pyautogui.press('tab')
    time.sleep(1)
    pyautogui.press('enter')
    print("DL......................")
    time.sleep(1)

    if i == 1:
        pyautogui.keyDown('shift')
        for _ in range(2):
            pyautogui.press('tab')
        pyautogui.keyUp('shift')
    pyautogui.press('enter')
    print("-------------------再検索-------------------")

# func1: 特定の国コードを指定して検索し、CSVをダウンロードする
def func1(n, data):
    # 検索条件入力
    pyautogui.sleep(6)
    pyautogui.press('right')
    for _ in range(2):
        pyautogui.press('tab')
    time.sleep(1)
    pyautogui.press('space')
    time.sleep(1)
    for _ in range(5):
        pyautogui.press('down')
    time.sleep(1)
    pyautogui.press('enter')
    for _ in range(3):
        pyautogui.press('tab')
    pyautogui.press('space')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.press('tab')
    time.sleep(1)
    pyautogui.typewrite('105')
    for _ in range(3):
        pyautogui.press('tab')
    pyautogui.press('space')
    pyautogui.press('down')
    pyautogui.press('enter')
    pyautogui.press('tab')
    pyautogui.typewrite(n)
    pyautogui.press('enter')
    print("TestFinished")
    output(data, 1)

# func2: 品目コードを指定して検索し、CSVをダウンロードする
def func2(n, data):
    time.sleep(4)
    print("Select code...........")
    for _ in range(10):
        pyautogui.press('tab')
    for _ in range(12):
        pyautogui.press('backspace')
    pyautogui.typewrite(n)
    pyautogui.press('enter')
    print("goto OutputPage")
    output(data, 2)

try:
    # ホームディレクトリまでのパスを取得
    p_path = os.path.expanduser("~")
    remove_glob(p_path + '/Downloads/*.zip')
    print("Removing DownloadFolder.......")

    x, y = pyautogui.position()

    # exebrowser関数の実行
    dldata(data)

    # ダウンロードファイルを一時フォルダに解凍し、リネームする
    today = datetime.date.today()  # 今日の日付を取得
    Y = str(today.year)
    M = str(today.month)
    D = str(today.day)

    # Path形式でDownloads以下のzipファイルを l_1にlistで格納
    p = Path(p_path, "Downloads/")
    l_1 = list(p.glob('*.zip'))
    re_l = ['fl.csv', 'fa.csv']

    Temp_P = Path(p_path, "DESKTOP/temp_python")

    # listの要素数から、zipfile.ZipFileリストをmyzipに格納
    for i in range(len(l_1)):
        with zipfile.ZipFile(str(l_1[i])) as myzip:
            # Temp_Pフォルダへ取り出し
            myzip.extractall(Temp_P)

    # temp_pythonフォルダの中のファイルを取得し、リネームして保存する
    files = sorted(Temp_P.glob('*.csv'))
    i = 0

    # 格納先フォルダの作成
    NEW_P = Path(p_path, "DESKTOP/temp_python/RENAME(PY_AUTO)/")
    if not os.path.exists(NEW_P):
        os.mkdir(NEW_P)
    else:
        shutil.rmtree(NEW_P)
        time.sleep(1)
        os.mkdir(NEW_P)

    # ファイルのリネーム
    for file in files:
        # 隠しファイルは処理しない
        if file.name.startswith('.'):
            continue

        # old Path
        oldFilePath = file.__str__()
        print("oldFilePath:", oldFilePath)

        # new Path
        newFileName = re_l[i]
        newFilePath = str(NEW_P) + '/' + newFileName
        print("newFileName:", newFileName)
        print("newFilePath:", newFilePath)

        print("[debug]",files)

        os.rename(oldFilePath, newFilePath)
        i += 1

    # CSVをEXCELに変換し、データをコピー＆ペーストする
    for csv_file in re_l:
        csv_path = os.path.join(p_path, f'Desktop/temp_python/RENAME(PY_AUTO)/{csv_file}')
        print("[debug]",csv_path)
        csv_file_path = os.path.join(p_path, 'Desktop/temp_python/RENAME(PY_AUTO)', csv_file)
        with open(csv_file_path, 'r', encoding='shift-jis', newline='') as tempfile:
            data_list_fl = list(csv.reader(tempfile, delimiter=','))

        wb = px.Workbook()
        ws = wb.active
        ws.append([])  # A1, B1 空行

        # データの格納
        for record in data_list_fl:
            ws.append(record)

        # xlsx形式で保存
        xlsx_file = os.path.splitext(csv_file)[0] + '.xlsx'
        xlsx_path = os.path.join(p_path, f'Desktop/temp_python/RENAME(PY_AUTO)/{xlsx_file}')
        wb.save(xlsx_path)

    # データを別のブックにコピー＆ペースト
    temp = os.path.join(p_path, 'Desktop/temp_python/RENAME(PY_AUTO)/temp.xlsx')
    output = os.path.join(p_path, 'Desktop/test.xlsx')

    wb_temp = px.load_workbook(filename=temp)
    ws_temp = wb_temp.active

    wb_output = px.Workbook()
    ws_output = wb_output.active

    for row in ws_temp.iter_rows(values_only=True):
        ws_output.append(row)

    # データを保存
    wb_output.save(output)

except KeyboardInterrupt:
    print('\n終了')
except Exception as e:
    print('エラーが発生しました:', str(e))
    # エラーメッセージを表示したり、適切な処理を行ったりする
