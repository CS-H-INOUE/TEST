import openpyxl
from openpyxl.utils import get_column_letter

def evaluate_excel_file(filename):
    workbook = openpyxl.load_workbook(filename, data_only=True)

    sheet_values = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        values = []
        for row in range(1, max_row + 1):
            row_values = []
            for column in range(1, max_column + 1):
                cell = sheet.cell(row=row, column=column)
                if cell.data_type == "f":
                    # 数式の評価
                    value = cell.value
                    value = sheet.cell(row=row, column=column).value
                    row_values.append(value)
                else:
                    row_values.append(cell.value)
            values.append(row_values)
        sheet_values[sheet_name] = values

    return sheet_values

# Excelファイルの読み込みとシートの値の評価
filename = "t.xlsx"
sheet_values = evaluate_excel_file(filename)

# シートの値を表示してみる（デバッグ用）
for sheet_name, values in sheet_values.items():
    print(f"Sheet: {sheet_name}")
    for row in values:
        print(row)
    print()
