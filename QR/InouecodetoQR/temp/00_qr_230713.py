import tkinter as tk
from tkinter import messagebox, font
import openpyxl
import os
import os.path
import qrcode
from datetime import date
from PIL import Image, ImageDraw, ImageFont









#
#
# ここから、商品の選定及びQRコードの生成
#
#

def load_excel_file(excel_file, sheet_name):
    """指定されたエクセルファイルとシートを読み込みます。"""
    print("[debug] load_excel_file")
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook[sheet_name]
    return sheet


def get_product_list(sheet):
    """指定されたシートから商品リストを取得します。"""
    print("[debug] get_product_list")
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            product = str(row[0]) + " " + str(row[1]) + " " + str(row[2]) + " " + str(row[4])
            products.append(product)
    return products


def search_products(products, search_input):
    """指定された商品リストから部分一致する商品を検索します。"""
    print("[debug] search_products")
    matching_products = []
    for product in products:
        if search_input.lower() in product.lower():
            matching_products.append(product)
    return matching_products


def show_matching_products(matching_products, qr_directory):
    """部分一致する商品の一覧を表示します。"""
    print("[debug] show_matching_products")
    window = tk.Toplevel()
    window.title("検索結果")

    # ウィンドウの幅と高さ
    window_width = 800
    window_height = 800

    # 画面の幅と高さを取得
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # ウィンドウを画面の中央に配置
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # メイリオフォントを使用するための設定
    font_name = "Meiryo"
    bold_font = font.Font(family=font_name, size=16, weight="bold")

    result_label = tk.Label(window, text="部分一致する商品:", font=bold_font)
    result_label.pack(pady=10)

    listbox = tk.Listbox(window, font=(font_name, 12))
    listbox.pack(fill=tk.BOTH, expand=True)

    for product in matching_products:
        listbox.insert(tk.END, product)

    def select_product():
        """商品の選択を行います。"""
        selected_index = listbox.curselection()
        if selected_index:
            selected_product = matching_products[selected_index[0]]
            generate_qr_code(selected_product, qr_directory)
            print(selected_product,"is slected!!!")
            messagebox.showinfo("商品選択", "選択された商品が保存されました。")

            # 選択された商品に対してQRコードを生成
            print("[debug: selected_product", selected_product)
            generate_qr_code(selected_product, qr_directory)

            window.destroy()
            return selected_product

        else:
            messagebox.showinfo("エラー", "商品が選択されていません。")
            window.destroy()

    # Enterキーを検索ボタンとしてバインド
    window.bind("<Return>", lambda event: select_product())

    select_button = tk.Button(window, text="選択", command=select_product, font=bold_font)
    select_button.pack(pady=10)

    window.mainloop()

def generate_qr_code(qr_text, qr_directory):
    """
    テキストを使用してQRコードを生成して保存する関数

    Args:
        qr_text (str): QRコードのテキスト
        qr_directory (str): QRコードの保存先ディレクトリ名

    Returns:
        None
    """
    print("[debug] generate_qr_code")
    # QRコードの保存先ディレクトリのパスを作成
    qr_directory_path = os.path.join(os.getcwd(), qr_directory)
    # QRコードの保存先ディレクトリが存在しない場合は作成
    if not os.path.exists(qr_directory_path):
        os.makedirs(qr_directory_path)

    # QRコードを生成
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_text)
    qr.make(fit=True)

    # QRコードをPIL Imageオブジェクトに変換
    qr_image = qr.make_image(fill_color="black", back_color="white")

    # QRコードのファイル名を作成
    qr_filename = f"qr_{qr_text}.png"

    # QRコードを保存
    qr_filepath = os.path.join(qr_directory_path, qr_filename)
    qr_image.save(qr_filepath)

def main():
    file_path = "data.xlsx"
    # ファイルが存在しない場合のみ作成
    if not os.path.isfile(file_path):
        # 新規に空のExcelファイルを作成
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    # エクセルファイルのパスとシート名を指定
    excel_file = 'codedata.xlsx'
    sheet_name = 'Sheet1'

    # QRコードの保存先ディレクトリを指定
    qr_directory = 'QRディレクトリ'

    # エクセルファイルを読み込む
    sheet = load_excel_file(excel_file, sheet_name)

    # 商品の検索と選択を行うGUIウィンドウを作成
    window = tk.Tk()
    window.title("商品検索")

    # ウィンドウの幅と高さ
    window_width = 500
    window_height = 400

    # 画面の幅と高さを取得
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # ウィンドウを画面の中央に配置
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    window.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # メイリオフォントを使用するための設定
    font_name = "Meiryo"
    bold_font = font.Font(family=font_name, size=16, weight="bold")
    # テキストを検索フィールドの前に表示
    label = tk.Label(window, text="検索したい文字列を入力してください.", font=("Meiryo", 14))
    label.pack(pady=10)
    label = tk.Label(window, text="（例. 50x 、T50、など）※半角数字で入力NG:５ OK:5", font=("Meiryo", 12))
    label.pack(pady=10)

    # 検索キーワードの入力フィールド
    search_entry = tk.Entry(window, font=("Meiryo", 16, "bold"))
    search_entry.pack(pady=10, ipadx=20, ipady=20)

    def search_button_click():
        """検索ボタンがクリックされたか、Enterキーが押されたときの処理を行います。"""
        search_input = search_entry.get().strip()

        if search_input:
            # 商品の一覧を取得
            products = get_product_list(sheet)

            # 部分一致する商品を検索
            matching_products = search_products(products, search_input)
            print("[debug: matching products" ,matching_products)

            if matching_products:
                # 部分一致する商品の一覧を表示して、QR作成
                show_matching_products(matching_products, qr_directory)

                messagebox.showinfo("完了", "選択された商品のQRコードが生成されました。")


                # 
                # 
                # ここから、次のステップへ
                # 
                # 


            else:
                messagebox.showinfo("検索結果", "部分一致する商品はありません。")


            # 再度検索を行う
            search_entry.delete(0, tk.END)  # 検索キーワードの入力フィールドをクリア
        else:
            messagebox.showinfo("エラー", "検索キーワードを入力してください。")

    def on_closing():
        """ウィンドウが閉じられるときの処理を行います。"""
        if messagebox.askokcancel("終了", "本当に終了しますか？"):
            window.destroy()

    # Enterキーを検索ボタンとしてバインド
    search_entry.bind("<Return>", lambda event: search_button_click())
    # 終了ボタン
    exit_button = tk.Button(window, text="終了", command=on_closing, font=bold_font)
    exit_button.pack(side=tk.BOTTOM, pady=10)

    # 検索ボタン
    search_button = tk.Button(window, text="検索", command=search_button_click, font=bold_font)
    search_button.pack(side=tk.BOTTOM, pady=10)

    # ウィンドウの"×"ボタンを押したときの処理を設定
    window.protocol("WM_DELETE_WINDOW", on_closing)

    # GUIのメインループを開始
    window.mainloop()


if __name__ == "__main__":
    main()
