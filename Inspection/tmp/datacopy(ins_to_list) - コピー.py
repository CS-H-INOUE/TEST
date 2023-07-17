import openpyxl
from openpyxl import calculation

# デバッグ関数
def debug_print(message):
    print("[DEBUG]:", message)

# inspection.xlsxの読み込み
debug_print("inspection.xlsxを読み込んでいます...")
inspection_wb = openpyxl.load_workbook('t.xlsx')

# list.xlsxの読み込み
debug_print("list.xlsxを読み込んでいます...")
list_wb = openpyxl.load_workbook('list.xlsx')

# inspection.xlsxの各シートに対して処理を行う
for sheet_name in inspection_wb.sheetnames:
    debug_print(f"シート '{sheet_name}' の処理を開始します...")
    inspection_sheet = inspection_wb[sheet_name]
    list_sheet = list_wb['Sheet1']

    # list.xlsxのC列の4行目以降のすべての値と比較（スペースを除去して比較）
    ''''''
    list_c_column = [str(cell.value).strip() for cell in list_sheet['C'][3:]]
    if sheet_name.strip() in list_c_column:
        last_matched_row = None  # 最後に一致した行の行数を格納する変数
        for i in range(11, inspection_sheet.max_row + 1):
            if inspection_sheet[f'C{i}'].value != None:
                c_value = inspection_sheet[f'C{i}'].value
                last_matched_row = i  # 最後に一致した行の行数を更新
                break

        debug_print(f"C{last_matched_row} の値が1以上です。B列の値をコピーします。")
        debug_print(f"Cvalue:{c_value}{inspection_sheet[f'B{last_matched_row}'].value}")
        list_sheet[f'M{i-3}'].value = inspection_sheet[f'B{last_matched_row}'].value

        if last_matched_row is not None:
            debug_print(f"最後に一致した行の行数: {last_matched_row}")
    ''''''

    # list.xlsxのC列の4行目以降のすべての値と比較（スペースを除去して比較）
    list_c_column = [str(cell.value).strip() for cell in list_sheet['C'][3:]]
    if sheet_name.strip() in list_c_column:
        last_matched_row = None  # 最後に一致した行の行数を格納する変数
        for idx, cell in enumerate(inspection_sheet['C'][10:], start=11):
            if cell.value is not None:
                c_value = cell.value
                last_matched_row = idx  # 最後に一致した行の行数を更新
                debug_print(f"一致した行番号: {idx - 3}")
                break

        b_value =  inspection_sheet[f'B{last_matched_row}'].value
        debug_print(f"C{last_matched_row} の値が1以上です。B列の値をコピーします。")
        debug_print("Cvalue", c_value, inspection_sheet[f'B{last_matched_row}'].value)
        mno = last_matched_row - 3  # i-3の値をmnoに格納
        list_sheet[f'M{mno}'].value = inspection_sheet[f'B{last_matched_row}'].value


# 結果をlist.xlsxとして保存
debug_print("結果を list.xlsx として保存します...")
list_wb.save('list.xlsx')

debug_print("処理が完了しました。")
