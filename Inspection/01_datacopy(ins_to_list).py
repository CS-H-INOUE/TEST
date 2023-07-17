'''
t.xlsxは値で保存する。
これを実行すると、検査元データからlistへ格納される
'''

import openpyxl

# デバッグ関数
def debug_print(message):
    print("[DEBUG]:", message)

# inspection.xlsxの読み込み
debug_print("inspection.xlsxを読み込んでいます...")
inspection_wb = openpyxl.load_workbook('t.xlsx')

# list.xlsxの読み込み
debug_print("list.xlsxを読み込んでいます...")
list_wb = openpyxl.load_workbook('list.xlsx')

# inspection.xlsxの各シートに対して処理を行う
for sheet_name in inspection_wb.sheetnames:
    debug_print(f"シート '{sheet_name}' の処理を開始します...")
    inspection_sheet = inspection_wb[sheet_name]
    list_sheet = list_wb['Sheet1']

    # list.xlsxのC列の4行目以降のすべての値と比較（スペースを除去して比較）
    list_c_column = [str(cell.value).strip() for cell in list_sheet['C'][3:]]
    if sheet_name.strip() in list_c_column:
        mno = list_c_column.index(sheet_name.strip())
        print("mno: ",mno)
        last_matched_row = None  # 最後に一致した行の行数を格納する変数
        for idx, cell in enumerate(inspection_sheet['C'][10:], start=11):
            if cell.value is not None and cell.value >1:
                c_value = cell.value
                last_matched_row = idx  # 最後に一致した行の行数を更新
                debug_print(f"一致した行番号: {idx}")

        # B列の値を取得
        b_value = inspection_sheet[f'B{last_matched_row}'].value
        debug_print(f"C{last_matched_row} の値が1以上です。B列の値をコピーします。")
        debug_print(f"Cvalue: {c_value}, b_value: {b_value}")
        debug_print(f"mno補正 +4: {mno+4}")
        list_sheet[f'M{mno+4}'].value = b_value

# 結果をlist.xlsxとして保存
debug_print("結果を list.xlsx として保存します...")
list_wb.save('list.xlsx')

debug_print("処理が完了しました。")
